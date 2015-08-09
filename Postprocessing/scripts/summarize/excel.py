import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Color, Fill, fills, PatternFill, Font

result_colors = {'perfect':'1F7218',
                 'almost':'68BE60',
                 'incomplete':'D1CB87',
                 'lowcov':'CCCCCC',
                 'errors':'E49999',
                 'dips':'84CCC9',
                 'nocall':'999999'
                }
text_colors =   {'perfect':'FFFFFF',
                 'almost':'FFFFFF',
                 'incomplete':'A53E3E',
                 'lowcov':'A53E3E',
                 'errors':'A53E3E',
                 'dips':'233D69',
                 'nocall':'CCCCCC'
                }

def create_result_workbook(refs,pools,calltable,bestbets,outfile):
  wb = Workbook()
  ws = wb.active

  ws.title = "bestbets"
  ws.append(["clone","best pool","variant","univF - reverse primer","forward - univR primer"])
  row = 2
  for ref in refs:
    ws['A%d' % row] = ref
    if bestbets[ref] is not None:
      c = calltable[ref][bestbets[ref]]
      ws['B%d' % row] = bestbets[ref]
      
      cellFill = PatternFill(start_color=result_colors[c['call']], fill_type=fills.FILL_SOLID) 
      ws['B%d' % row].fill = cellFill
      if c['call'] == 'almost':
        ws['D%d' % row] = c['p1']
        ws['E%d' % row] = c['p2']
        for col in ['C','D','E']:
          cellFill = PatternFill(start_color=result_colors[c['call']], fill_type=fills.FILL_SOLID) 
          ws['%s%d' % (col,row)].fill = cellFill
    row += 1
  
  ws = wb.create_sheet()
  ws.title = 'overview'
  ws.append([""] + pools)
  cols = dict((p,chr(66+i)) for i,p in enumerate(pools))
  row = 2
  for ref in refs:
    ws['A%d' % row] = ref
    for p in pools:
      c = calltable[ref][p] #['call']
      ws['%s%d' % (cols[p],row)] = c['call']

      cellFill = PatternFill(start_color=result_colors[c['call']], fill_type=fills.FILL_SOLID)
      cellFont = Font(color=text_colors[c['call']])
      ws['%s%d' % (cols[p],row)].fill = cellFill
      ws['%s%d' % (cols[p],row)].font = cellFont
    row += 1
  
  for pool in pools:
    ws = wb.create_sheet()
    ws.title = pool
    ws.append(["clone","result","variant","univF - reverse primer","forward - univR primer"])
    row = 2
    for ref in refs:
      ws['A%d' % row] = ref
      c = calltable[ref][pool]
      ws['B%d' % row] = c['call']

      cellFill = PatternFill(start_color=result_colors[c['call']], fill_type=fills.FILL_SOLID)
      cellFont = Font(color=text_colors[c['call']])
      ws['B%d' % row].fill = cellFill
      ws['B%d' % row].font = cellFont
      if c['call'] == 'almost':
        ws['D%d' % row] = c['p1']
        ws['E%d' % row] = c['p2']
        for col in ['C','D','E']:
          cellFill = PatternFill(start_color=result_colors[c['call']], fill_type=fills.FILL_SOLID)
          cellFont = Font(color=text_colors[c['call']])
          ws['%s%d' % (col,row)].fill = cellFill
          ws['%s%d' % (col,row)].font = cellFont
      row += 1
  
  wb.save(filename=outfile)  
